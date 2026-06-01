from openpyxl import Workbook # tạo file Excel mới
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter # đổi số cột sang chữ cột Excel. 1 -> A

#Font         → chỉnh chữ
#Alignment    → căn lề
#PatternFill  → tô màu nền ô
#Border       → tạo viền ô
#Side         → kiểu đường viền


NOT_AVAILABLE = "N/A"


# tạo tiêu đề lớn ở dòng 1 của mỗi sheet
def set_title(ws, title, end_col):
    # Gộp ô tiêu đề: A1:E1
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_col
    )

    # Lấy ô A1, nội dung sẽ được đặt ở ô đầu tiên là A1
    cell = ws.cell(row=1, column=1)
    cell.value = title # Gán tiêu đề
    # Định dạng chữ tiêu đề
    cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )
    # Tô nền tiêu đề màu xanh đậm
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )
    # Căn giữa tiêu đề
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Chỉnh chiều cao dòng 1
    ws.row_dimensions[1].height = 28


# Format dòng header của bảng Excel
def style_header(row):
    # Duyệt từng ô trong dòng
    for cell in row:
        # Mỗi ô header sẽ có Chữ đậm, Màu trắng
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        # Tô nền
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="5B9BD5" # xanh dương
        )
        # Căn lề
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


# Kẻ viền toàn bộ sheet
def apply_border(ws):
    # Tạo kiểu viền mảnh
    thin = Side(
        border_style="thin",
        color="BFBFBF"
    )

    # Tạo border hoàn chỉnh
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # Duyệt toàn bộ sheet
    for row in ws.iter_rows():
        # Duyệt từng ô
        for cell in row:
            cell.border = border # Gán viền
            # Căn lề
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True # Căn lề
            )


# tự động chỉnh độ rộng cột trong Excel
def auto_width(ws, max_width=65):
    # Duyệt từng cột trong sheet
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column) # Lấy chữ cái của cột
        max_len = 0 # Khởi tạo độ dài lớn nhất

        # Duyệt từng ô trong cột
        for cell in col:
            # Kiểm tra ô có dữ liệu không
            if cell.value is not None:
                # Tính độ dài lớn nhất
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )

        # Gán độ rộng cột
        ws.column_dimensions[col_letter].width = min(
            max_len + 4,  # nd + khoảng trống
            max_width
        )


# trả về màu nền cho ô Excel theo điểm số
def get_score_fill(value):
    # Thử ép value sang số
    try:
        value = float(value)
    # Nếu không ép được -> xanh xám nhạt
    except Exception:
        return PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

    # Nếu điểm từ 90 trở lên thì tô màu xanh
    if value >= 90:
        return PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )

    # Nếu điểm từ 80 đến dưới 90 thì tô màu vàng
    if value >= 80:
        return PatternFill(
            fill_type="solid",
            fgColor="FFEB9C"
        )

    # Nếu điểm dưới 80 thì tô màu đỏ nhạt
    return PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )


# format một ô trong cột Value của sheet Summary
def format_percent_or_text(cell):
    # Trường hợp ô là N/A
    if cell.value == NOT_AVAILABLE:
        # Tô màu xanh xám nhạt
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )
        # Chữ in đậm, màu xám
        cell.font = Font(
            bold=True,
            color="666666"
        )
        return # dừng hàm

    # Trường hợp ô là số
    try:
        value = float(cell.value)
        cell.fill = get_score_fill(value) # Tô màu theo điểm
        cell.value = value / 100 # Chuyển số thành dạng phần trăm trong Excel
        cell.number_format = "0.00%" # Format hiển thị phần trăm
    # Nếu không phải số và cũng không phải N/A
    except Exception:
        pass # bỏ qua


# tạo sheet đầu tiên chứa mô tả phương pháp đánh giá
def create_method_sheet(wb, function_name, web_name):
    ws = wb.active # Tạo sheet đầu tiên
    ws.title = "Evaluation_Method" # Đổi tên sheet

    # Tạo tiêu đề
    set_title(
        ws,
        "AUTO TESTING EVALUATION MODEL",
        4
    )

    # Thêm dòng trống
    ws.append([])
    ws.append(["Function", function_name]) # thêm function
    ws.append(["Website", web_name])
    ws.append(["Evaluation Target", "AI-generated test cases and keyword test scripts"]) # Thêm đối tượng đánh giá
    ws.append(["Evaluation Type", "Automatic metrics-based evaluation"]) # Thêm loại đánh giá

    ws.append([])
    # Thêm header bảng metric
    ws.append([
        "Evaluation Layer",
        "Metric",
        "Formula",
        "Meaning"
    ])

    # Tạo danh sách rows
    rows = [
        [
            "Test Case Quality",
            "Execution Path Coverage",
            "Paths with Test Case / Total Execution Paths",
            "Đo AI có sinh đủ test case cho các execution path hay không"
        ],
        [
            "Test Case Quality",
            "Step Completeness",
            "Preserved User Actions / Required User Actions",
            "Đo test case có giữ đủ bước user action quan trọng không"
        ],
        [
            "Test Case Quality",
            "Expected Result Accuracy",
            "Matched Expected Results / Required Expected Results",
            "Đo expected_result có khớp execution path không"
        ],
        [
            "Test Case Quality",
            "Validation Pass Rate",
            "Validation PASS Test Cases / Total Test Cases",
            "Đo tỷ lệ test case pass validator"
        ],
        [
            "Test Script Verification",
            "Script Generation Coverage",
            "Generated Scripts / Total Test Cases",
            "Đo test case có chuyển được thành keyword script không"
        ],
        [
            "Test Script Verification",
            "Script Execution Pass Rate",
            "Passed Scripts / Executed Scripts",
            "Chỉ tính khi đã có log thực thi test script"
        ]
    ]

    # Ghi rows vào Excel
    for row in rows:
        ws.append(row)

    style_header(ws[7]) # Format header
    apply_border(ws) # Kẻ viền
    auto_width(ws) # Tự động chỉnh độ rộng
    ws.freeze_panes = "A8" # Khóa các dòng phía trên A8

    return ws


# tạo sheet tên Summary để ghi các chỉ số tổng hợ
def create_summary_sheet(wb, summary):
    ws = wb.create_sheet("Summary") # Tạo sheet mới

    # Tạo tiêu đề sheet từ A1:E1
    set_title(
        ws,
        "EVALUATION SUMMARY",
        5
    )

    ws.append([])
    # Thêm header của bảng
    ws.append([
        "Group",
        "Metric",
        "Formula",
        "Value",
        "Note"
    ])

    # Duyệt từng metric trong summary
    for item in summary:
        # Ghi từng metric vào Excel
        ws.append([
            item.get("group", ""),
            item.get("metric", ""),
            item.get("formula", ""),
            item.get("value", ""),
            item.get("note", "")
        ])

    style_header(ws[3]) # Format header

    # Format cột Value
    # duyệt từ dòng 4 đến dòng cuối cùng
    for row in range(4, ws.max_row + 1):
        format_percent_or_text(
            ws.cell(row=row, column=4)
        )

    apply_border(ws) # Kẻ viền toàn sheet
    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


# tạo sheet TestCase_Evaluation ghi chi tiết từng execution path/test case.
def create_testcase_sheet(wb, testcase_details):
    ws = wb.create_sheet("TestCase_Evaluation")

    set_title(
        ws,
        "TEST CASE QUALITY EVALUATION",
        10
    )

    ws.append([])
    # Thêm header
    ws.append([
        "Path ID",
        "Test Case ID",
        "Path Covered",
        "Required Actions",
        "Preserved Actions",
        "Missing Actions",
        "Expected Count",
        "Expected Matched",
        "Missing Expected",
        "Validation Status"
    ])

    for item in testcase_details:
        # Ghi từng dòng vào Excel
        ws.append([
            item.get("path_id", ""),
            item.get("test_case_id", ""),
            "Yes" if item.get("path_covered") else "No", # Nếu path có test case thì ghi: Yes
            item.get("required_actions", 0), # Ghi số bước user action bắt buộc phải có
            item.get("preserved_actions", 0), # Ghi số bước user action mà test case giữ lại được
            "\n".join(item.get("missing_actions", [])), # Ghi các bước bị thiếu
            item.get("expected_count", 0), # Ghi số expected result cần kiểm tra
            item.get("expected_matched", 0), # Ghi số expected result đã khớp
            "\n".join(item.get("missing_expected", [])), # Ghi các expected result bị thiếu
            item.get("validation_status", "") # Ghi trạng thái validator
        ])

    style_header(ws[3])
    apply_border(ws)
    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


# Tạo sheet ghi kết quả đánh giá phần test script
def create_script_sheet(wb, script_details):
    ws = wb.create_sheet("TestScript_Evaluation")

    set_title(
        ws,
        "TEST SCRIPT VERIFICATION",
        5
    )

    ws.append([])
    # Thêm header
    ws.append([
        "Test Case ID",
        "Path ID",
        "Script Generated",
        "Execution Status",
        "Execution Error"
    ])

    # Ghi dữ liệu chi tiết từng script
    for item in script_details:
        # Ghi từng dòng vào Excel
        ws.append([
            item.get("test_case_id", ""),
            item.get("path_id", ""),
            "Yes" if item.get("script_generated") else "No",
            item.get("execution_status", NOT_AVAILABLE),
            item.get("execution_error", "")
        ])

    style_header(ws[3])
    apply_border(ws)
    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


# hàm chính của file export Excel
def export_auto_evaluation_to_excel(
        summary,
        testcase_details,
        script_details,
        output_path,
        function_name="",
        web_name=""
):
    wb = Workbook() # tạo một file Excel mới

    # Tạo sheet Evaluation_Method
    create_method_sheet(
        wb,
        function_name,
        web_name
    )

    # Tạo sheet Summary
    create_summary_sheet(
        wb,
        summary
    )

    # Tạo sheet TestCase_Evaluation
    create_testcase_sheet(
        wb,
        testcase_details
    )

    # Tạo sheet TestScript_Evaluation
    create_script_sheet(
        wb,
        script_details
    )

    wb.save(output_path) # Lưu workbook ra file