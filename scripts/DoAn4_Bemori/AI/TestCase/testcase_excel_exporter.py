import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy
from unidecode import unidecode


TESTER_NAME = "Lê Thị Ánh"

# Chuẩn hóa tên sheet: bỏ dấu TV, xoá khoảng trắng
def normalize_sheet_name(name: str):
    return unidecode(name).replace(" ", "")[:31]


# Cho phép xuống dòng trong cell
def apply_wrap_text(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")


# Auto height theo số dòng
def auto_adjust_row_height(ws, row):
    max_lines = 1

    for col in range(1, 6):  # A → E
        cell = ws.cell(row=row, column=col)

        if cell.value:
            lines = str(cell.value).count("\n") + 1
            if lines > max_lines:
                max_lines = lines

    ws.row_dimensions[row].height = max_lines * 15


# Set width cố định
def auto_adjust_column_width(ws):
    column_widths = {
        "A": 12, "B": 40, "C": 30,
        "D": 45, "E": 45, "F": 25,
        "G": 12, "H": 15, "I": 25
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


# Copy toàn bộ style từ dòng template (giữ nguyên format excel)
def copy_style(src, dest):
    dest.font = copy(src.font)
    dest.border = copy(src.border)
    dest.fill = copy(src.fill)
    dest.number_format = src.number_format
    dest.alignment = copy(src.alignment)


# Update số lượng test case
def update_case_count(ws, count):
    ws["E3"] = f"Number of cases: {count}"
    ws["F3"] = f"Number of cases: {count}"


# Ghi test case vào Excel theo format chuẩn dựa trên template có sẵn
def export_testcases_to_excel(testcases, module_name, template_path):
    wb = load_workbook(template_path)

    sheet_name = normalize_sheet_name(module_name)

    # Nếu sheet đã tồn tại → dùng lại
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

    # Nếu chưa có → copy từ template sheet
    else:
        template_ws = wb.active
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_name

    # Ghi: Tên module, Tên tester
    ws["B2"] = sheet_name
    ws["B3"] = TESTER_NAME

    # Xác định vị trí ghi data
    start_row = 5
    template_row = 5
    total_cases = len(testcases)

    # XOÁ DÒNG THỪA
    last_needed_row = start_row + total_cases - 1

    if ws.max_row > last_needed_row:
        ws.delete_rows(last_needed_row + 1, ws.max_row - last_needed_row)

    # THÊM DÒNG THIẾU
    needed_last_row = start_row + total_cases - 1

    if ws.max_row < needed_last_row:
        for _ in range(needed_last_row - ws.max_row):
            ws.append([""] * 9)

            new_row = ws.max_row
            for col in range(1, 10):
                copy_style(
                    ws.cell(row=template_row, column=col),
                    ws.cell(row=new_row, column=col)
                )

    # Ghi dữ liệu test case
    test_date = datetime.today().strftime("%d/%m/%Y")

    for i, tc in enumerate(testcases):
        row = start_row + i

        ws.cell(row=row, column=1, value=tc.get("test_case_id"))
        ws.cell(row=row, column=2, value=tc.get("scenario"))
        ws.cell(row=row, column=3, value="\n".join(tc.get("precondition", [])))
        ws.cell(row=row, column=4, value="\n".join(tc.get("steps", [])))
        ws.cell(row=row, column=5, value="\n".join(tc.get("expected_result", [])))
        ws.cell(row=row, column=8, value=test_date)

        # Copy style
        for col in range(1, 10):
            copy_style(
                ws.cell(row=template_row, column=col),
                ws.cell(row=row, column=col)
            )

        # Wrap text
        for col in range(1, 6):
            apply_wrap_text(ws.cell(row=row, column=col))

        auto_adjust_row_height(ws, row)

    update_case_count(ws, total_cases)
    auto_adjust_column_width(ws)

    wb.save(template_path)