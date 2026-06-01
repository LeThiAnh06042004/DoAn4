import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


NOT_AVAILABLE = "N/A" # hằng số Dùng khi không có dữ liệu


# Lấy thư mục gốc project
def get_project_root():
    return os.path.abspath(
        os.path.join(
            os.path.dirname(__file__),
            "..",
            ".."
        )
    )


# Lấy đường dẫn file đánh giá Excel
def get_evaluation_file_path(function_name, web_name):
    project_root = get_project_root()

    return os.path.join(
        project_root,
        "AI",
        "TestCase",
        "Evaluation",
        f"EV_{function_name}_{web_name}.xlsx"
    )


# Hàm tính phần trăm
def percent(numerator, denominator):
    if denominator == 0:
        return NOT_AVAILABLE

    return round(
        (numerator / denominator) * 100,
        2
    )


# Tô màu trạng thái PASS/FAIL
def get_status_fill(status):
    if status == "PASS":
        return PatternFill(
            fill_type="solid",
            fgColor="C6EFCE" # màu xanh nhạt
        )

    if status == "FAIL":
        return PatternFill(
            fill_type="solid",
            fgColor="FFC7CE" # màu đỏ nhạt
        )

    return PatternFill(
        fill_type="solid",
        fgColor="D9EAD3" # màu xanh xám nhạt
    )


# tô màu theo điểm hoặc tỷ lệ phần trăm
def get_score_fill(value):
    try:
        value = float(value)
    except Exception:
        return PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

    if value >= 90:
        return PatternFill(
            fill_type="solid",
            fgColor="C6EFCE" # màu xanh, kết quả tốt
        )

    if value >= 80:
        return PatternFill(
            fill_type="solid",
            fgColor="FFEB9C" # màu vàng, mức khá
        )

    return PatternFill(
        fill_type="solid",
        fgColor="FFC7CE" # màu đỏ, kết quả chưa tốt
    )


# Cập nhật sheet Summary
def update_summary_sheet(wb, execution_results):
    ws = wb["Summary"]

    executed_count = len(execution_results) # Đếm số testcase đã chạy

    # Đếm số testcase PASS
    passed_count = sum(
        1 for item in execution_results
        if item.get("status") == "PASS"
    )

    # Tính pass rate
    pass_rate = percent(
        passed_count,
        executed_count
    )

    # Tìm dòng metric cần cập nhật
    for row in range(1, ws.max_row + 1):
        metric = ws.cell(row=row, column=2).value # Nó kiểm tra cột 2 xem có dòng nào là Script Execution Pass Rate

        # Tìm thấy: cập nhật ở cột 4, ghi chú ở cột 5
        if metric == "Script Execution Pass Rate":
            value_cell = ws.cell(row=row, column=4)
            note_cell = ws.cell(row=row, column=5)

            # Nếu chưa có kết quả thực thi thì ghi N/A
            if pass_rate == NOT_AVAILABLE:
                value_cell.value = NOT_AVAILABLE
                value_cell.fill = get_score_fill(NOT_AVAILABLE)
                value_cell.font = Font(
                    bold=True,
                    color="666666"
                )
                note_cell.value = "Chưa có kết quả thực thi test script" # note
            # Nếu có kết quả thực thi
            else:
                value_cell.value = pass_rate / 100
                value_cell.number_format = "0.00%"
                value_cell.fill = get_score_fill(pass_rate)
                note_cell.value = "Đã tự động cập nhật từ kết quả thực thi test script"

            break


# Cập nhật sheet TestScript_Evaluation
# cập nhật chi tiết từng test case trong sheet
def update_test_script_sheet(wb, execution_results):
    ws = wb["TestScript_Evaluation"]

    # Tạo map kết quả thực thi để tra cứu nhanh theo test_case_id
    execution_map = {
        item.get("test_case_id"): item
        for item in execution_results
    }

    # Duyệt từng dòng test case
    for row in range(4, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=1).value # Lấy mã test case

        execution = execution_map.get(tc_id) # Lấy kết quả tương ứng

        # Ghi trạng thái
        status_cell = ws.cell(row=row, column=4)
        error_cell = ws.cell(row=row, column=5)

        # Nếu có kết quả
        if execution:
            status = execution.get(
                "status",
                NOT_AVAILABLE
            )

            status_cell.value = status # ghi trạng thái
            status_cell.fill = get_status_fill(status)  # tô màu
            status_cell.font = Font(bold=True)

            # Nếu fail thì ghi lỗi vào cột 5
            error_cell.value = execution.get(
                "error",
                ""
            )
        # Nếu không có kết quả
        else:
            status_cell.value = NOT_AVAILABLE # tc đó có trong file Excel nhưng không được chạy trong lần này.
            status_cell.fill = get_status_fill(NOT_AVAILABLE)
            status_cell.font = Font(
                bold=True,
                color="666666"
            )
            error_cell.value = ""


# Hàm chính cập nhật evaluation
def update_evaluation_after_execution(
        function_name,
        web_name,
        execution_results
):
    # lấy file evaluation
    evaluation_path = get_evaluation_file_path(
        function_name,
        web_name
    )

    # Kiểm tra file có tồn tại không
    if not os.path.exists(evaluation_path):
        print(
            f"Không tìm thấy file đánh giá: {evaluation_path}"
        )
        return

    # Mở file Excel để cập nhật
    wb = load_workbook(
        evaluation_path
    )

    # Cập nhật sheet Summary
    update_summary_sheet(
        wb,
        execution_results
    )

    # Cập nhật sheet TestScript_Evaluation
    update_test_script_sheet(
        wb,
        execution_results
    )

    # Lưu file
    wb.save(
        evaluation_path
    )

    print("===== AUTO UPDATE EVALUATION DONE =====")
    print(f"Evaluation file: {evaluation_path}")
    print(f"Executed: {len(execution_results)}")
    print(
        "Passed:",
        sum(
            1 for item in execution_results
            if item.get("status") == "PASS"
        )
    )