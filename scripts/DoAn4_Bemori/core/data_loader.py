import os
import yaml
import json
import csv
import openpyxl

# Lấy thư mục gốc của project
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")


class DataLoader:
    @staticmethod
    def load_yaml(file_name):
        """Load file YAML trong thư mục data"""
        path = os.path.join(DATA_DIR, file_name)
        if not os.path.exists(path):
            raise FileNotFoundError(f"Không tìm thấy file YAML: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    @staticmethod
    def load_json(file_name):
        """Load file JSON trong thư mục data"""
        path = os.path.join(DATA_DIR, file_name)
        if not os.path.exists(path):
            raise FileNotFoundError(f"Không tìm thấy file JSON: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def load_csv(file_name):
        """Load file CSV trong thư mục data"""
        path = os.path.join(DATA_DIR, file_name)
        if not os.path.exists(path):
            raise FileNotFoundError(f"Không tìm thấy file CSV: {path}")
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return [row for row in reader]

    @staticmethod
    def load_xlsx(file_name, sheet_name=None):
        """
        Load file Excel (xlsx) trong thư mục data.
        Trả về list[dict] (mỗi row là một dict).
        """
        path = os.path.join(DATA_DIR, file_name)
        if not os.path.exists(path):
            raise FileNotFoundError(f"Không tìm thấy file XLSX: {path}")

        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        # Lấy header từ dòng đầu tiên
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        data = []

        # Lấy từng row -> dict
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        return data
